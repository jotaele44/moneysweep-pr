#!/usr/bin/env python3
"""Project the PR Federal Publications master into federation source rows.

Source: ``Puerto_Rico_Federal_Publications_Master_v7.xlsx`` ("All Publications"
sheet, 4,617 federal publications — DOE/NREL/OSTI/USGS/EPA reports with
Funder/Agency, Year, Report ID, URL). Produces ``data/sources/federal_publications.jsonl``
in the federation **source** schema (deterministic ``src_<32hex>`` ids, lineage,
synthetic=false), so the publications are available as evidence-grade federation
sources for the money-intelligence node.

This is additive and standalone — it does NOT modify the canonical_v1 bridge or
its committed tables. Source xlsx lives outside the repo; pass ``--src`` to override.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

import openpyxl

DEFAULT_SRC = "/Users/jotaele/Documents/Data/Puerto_Rico_Federal_Publications_Master_v7.xlsx"
SHEET = "All Publications"


def _sid(key: str) -> str:
    return "src_" + hashlib.sha256(str(key).encode()).hexdigest()[:32]


def _created(year) -> str:
    y = str(year or "").strip()[:4]
    return f"{y}-01-01T00:00:00Z" if y.isdigit() else "1970-01-01T00:00:00Z"


def build_sources(rows: list) -> list:
    out: dict = {}
    for r in rows:
        title = (r.get("Title") or "").strip()
        if not title:
            continue
        report_id = (r.get("Report ID") or "").strip()
        url = (r.get("URL") or "").strip()
        key = report_id or url or title
        created = _created(r.get("Year"))
        row = {
            "source_id": _sid(key),
            "source_type": (r.get("Doc Type") or "federal_publication").strip() or "federal_publication",
            "source_name": title[:300],
            "confidence": 1.0,
            "lineage": {
                "producer_script": "scripts/ingest_federal_publications.py",
                "producer_phase": "FEDERAL_PUBLICATIONS_INGEST",
                "source_inputs": ["Puerto_Rico_Federal_Publications_Master_v7.xlsx#All Publications"],
                "extraction_method": "xlsx_row_projection",
            },
            "synthetic": False,
            "created_at": created,
            "extracted_at": created,
        }
        if url.startswith("http"):
            row["source_url"] = url
        else:
            row["source_ref"] = str(key)
        out[row["source_id"]] = row  # dedupe by deterministic id
    return list(out.values())


def _read_sheet(path: Path) -> list:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[SHEET]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: v for h, v in zip(header, values) if h})
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--out", default="data/sources/federal_publications.jsonl")
    args = ap.parse_args()

    sources = build_sources(_read_sheet(Path(args.src)))
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("".join(json.dumps(s, sort_keys=True) + "\n" for s in sources))
    print(f"wrote {len(sources)} federation sources -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
